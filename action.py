# -*- coding: UTF-8 -*-
import requests
import re
import time
import openpyxl


def login(account, password):
    head = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36',
        'Referer': 'https://jksb.v.zzu.edu.cn/vls6sss/zzujksb.dll/first0?fun2=a',
        'Content-Type': 'application/x-www-form-urlencoded'
    }
    login_data = {
        'uid': account,
        'upw': password,
        'smbtn': '进入健康状况上报平台',
        'hh28': '969'
    }

    login_url = 'https://jksb.v.zzu.edu.cn/vls6sss/zzujksb.dll/login'
    login_response = requests.post(url=login_url, headers=head, data=login_data, proxies=proxy)
    ptopid = re.findall(re.compile('ptopid=(.*)&sid'), login_response.content.decode('utf-8'))
    sid = re.findall(re.compile('&sid=(.*)"'), login_response.content.decode('utf-8'))

    head = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:88.0) Gecko/20100101 Firefox/88.0',
        'Referer': 'https://jksb.v.zzu.edu.cn/vls6sss/zzujksb.dll/jksb?ptopid=' + ptopid[0] + '&sid=' + sid[0] + '&fun2=',
        'Content-Type': 'application/x-www-form-urlencoded'
    }
    update_data = {
        "day6": "b",
        "did": "1",
        "door": "",
        "men6": "a",
        "ptopid": ptopid[0],
        "sid": sid[0]
    }

    update_url = 'https://jksb.v.zzu.edu.cn/vls6sss/zzujksb.dll/jksb'
    update_response = requests.post(url=update_url, headers=head, data=update_data, proxies=proxy)
    sid = re.findall(re.compile('&sid=(.*)" marginwidth'), update_response.content.decode('utf-8'))

    return ptopid[0], sid[0]


def post(ptopid, sid):
    post_data = {
        'myvs_1': '否',
        'myvs_2': '否',
        'myvs_3': '否',
        'myvs_4': '否',
        'myvs_5': '否',
        'myvs_6': '否',
        'myvs_7': '否',
        'myvs_8': '否',
        'myvs_9': '否',
        'myvs_10': '否',
        'myvs_11': '否',
        'myvs_12': '否',
        'myvs_13': 'g',
        'myvs_13a': '41',               # 41是河南省
        'myvs_13b': '4101',             # 4101是郑州市
        'myvs_13c': 修改为具体地址,
        'myvs_24': '否',
        'myvs_26': "2",
        'memo22': '[待定]',
        'did': '2',
        'door': '',
        'day6': 'b',
        'men6': 'a',
        'sheng6': '',
        'shi6': '',
        'fun3': '',
        'jingdu': '',
        'weidu': '',
        'ptopid': ptopid,
        'sid': sid
    }
    head = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36',
        'Referer': 'https://jksb.v.zzu.edu.cn/vls6sss/zzujksb.dll/jksb',
        'Content-Type': 'application/x-www-form-urlencoded'
    }
    post_url = 'https://jksb.v.zzu.edu.cn/vls6sss/zzujksb.dll/jksb'
    requests.post(url=post_url, headers=head, data=post_data, proxies=proxy)


def if_success(ptopid):
    time.sleep(2)
    head = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36'}
    url = 'https://jksb.v.zzu.edu.cn/vls6sss/zzujksb.dll/jksb?ptopid='+ptopid
    response = requests.get(url=url, headers=head)

    if '今日您已经填报过了' in response.content.decode('utf-8'):
        return True
    else:
        return False


def send_message(success, uid):
    baseUrl = 'http://wxpusher.zjiecode.com/api/send/message/?appToken='
    appToken = WxPusher创建的应用的AppToken
    if success:
        content = '%E6%89%93%E5%8D%A1%E6%88%90%E5%8A%9F'
    else:
        content = '%E6%89%93%E5%8D%A1%E5%A4%B1%E8%B4%A5%EF%BC%8C%E8%AF%B7%E6%89%8B%E5%8A%A8%E6%89%93%E5%8D%A1'

    url = baseUrl + appToken + '&content=' + content + '&uid=' + uid
    requests.get(url)


# 这段是代理IP可以去天启ip注册一下 
# IP = requests.get('http://api.tianqiip.com/getip?secret=*********************&type=txt&num=1&time=3&region=410000&cs=1&port=2').text.rstrip()
# proxy = {'https': IP}

若不使用代理IP用这行
# proxy = {'https': None}

account_workbook = openpyxl.load_workbook('C:/script/account.xlsx')  修改这里路径，最好是绝对路径
account_sheet = account_workbook.active

count = 2
todo_list = list(range(1, account_sheet.max_row + 1))

while True:
    count -= 1
    index = 0
    while index < len(todo_list):
        time.sleep(2)
        temp_account = account_sheet.cell(todo_list[index], 1).value
        temp_password = account_sheet.cell(todo_list[index], 2).value
        temp_uid = account_sheet.cell(todo_list[index], 3).value

        temp_ptopid, temp_sid = login(temp_account, temp_password)
        post(temp_ptopid, temp_sid)

        if if_success(temp_ptopid):
            send_message(True, temp_uid)
            todo_list.remove(todo_list[index])
        else:
            index += 1

    if count == 0:
        break
    if not todo_list:
        break

for i in todo_list:
    send_message(False, account_sheet.cell(i, 3).value)
